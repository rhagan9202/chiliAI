"""XLSX parser."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from ingestion.models import DocumentFormat, ParsedDocument, SourceDocument, StructuredRecord
from ingestion.parsers.exceptions import ParserError
from ingestion.parsers.utils import build_parser_metadata


class XlsxParser:
    """Parse Excel workbooks into structured records."""

    name = "xlsx"
    version = "1.0"
    supported_formats = (DocumentFormat.XLSX,)

    def parse(self, source: SourceDocument, content: bytes) -> ParsedDocument:
        try:
            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParserError(f"Unable to read XLSX workbook: {exc}") from exc

        records: list[StructuredRecord] = []
        sheet_names: list[str] = []
        record_index = 0

        for worksheet in workbook.worksheets:
            sheet_names.append(worksheet.title)
            row_iter = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(row_iter)
            except StopIteration:
                continue

            headers = [
                str(value).strip() if value not in (None, "") else f"column_{index + 1}"
                for index, value in enumerate(header_row)
            ]

            for row_number, row in enumerate(row_iter, start=1):
                if all(value in (None, "") for value in row):
                    continue
                fields = {headers[index]: row[index] for index in range(len(headers))}
                records.append(
                    StructuredRecord(
                        id=f"{source.id}-row-{record_index}",
                        row_number=row_number,
                        fields=fields,
                        metadata={"sheet_name": worksheet.title},
                    )
                )
                record_index += 1

        if not records:
            raise ParserError("XLSX workbook does not contain any data rows.")

        return ParsedDocument(
            id=f"parsed-{source.id}",
            source_document_id=source.id,
            records=records,
            parser_name=self.name,
            parser_version=self.version,
            parser_metadata=build_parser_metadata(
                workbook_sheets=sheet_names,
                row_count=len(records),
            ),
        )